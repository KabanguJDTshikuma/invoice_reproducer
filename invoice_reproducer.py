"""Reproduced invoice based on total amount of the item's price"""
import openpyxl
import random
import xlsxwriter


wb = openpyxl.load_workbook('PRICE LIST.xlsx')
sheetList = wb.get_sheet_by_name('Sheet1')
sheetTotal = wb.get_sheet_by_name('Sheet2')

total_amount = {}
product_dict = {}
price_list = []
somme = {}
amount = 0.0
item_list = []
product_list = []

# construct of dictionary of all the product(key) unit price(value)

for i in range(2, 143):#row_item):
    if sheetList.cell(row=i, column=2).value != None:
        Item = sheetList.cell(row=i, column=1).value
        Item_weight = sheetList.cell(row=i, column=2).value
        product = Item + ': ' + Item_weight

        product_dict[product] = sheetList.cell(row=i, column=3).value
        single_product = (product, sheetList.cell(row=i, column=3).value)
        if product_dict[product] != None:
            price_list.append(sheetList.cell(row=i, column=3).value)
            product_list.append(single_product)

    else:
        Item = sheetList.cell(row=i, column=1).value
        Item_weight = sheetList.cell(row=i, column=2).value
        product = Item
        product_dict[Item] = sheetList.cell(row=i, column=3).value
        single_product = (product, sheetList.cell(row=i, column=3).value)
        if product_dict[product] != None:
            price_list.append(sheetList.cell(row=i, column=3).value)
            product_list.append(single_product)


print("Which number of row do you want to start:")
row_startNum = int(input(">"))
print("How many invoice do you want to print")
last_row = int(input(">"))

sheetdate = []
n_row = 1
#row_startNum = 1 # write here the Beginning of the row
r = 3  # Don't touch
date_row = 0
#while n_row <= 1: # write here the number of the last row
while n_row <= last_row:

    for i in range(row_startNum, row_startNum + 1):
        total_amount[sheetTotal.cell(row=i, column=1).value] = sheetTotal.cell(row=i, column=2).value
        sheetdate.append(sheetTotal.cell(row=i, column=1).value)
        while amount != total_amount[sheetTotal.cell(row=i, column=1).value]:
            product_choice = random.choice(product_list)
            # print(product_choice)
            item_list.append(product_choice)
            amount += float(product_choice[1])
            if amount > total_amount[sheetTotal.cell(row=i, column=1).value]:
                amount = 0.0
                item_list = []
            elif amount == total_amount[sheetTotal.cell(row=i, column=1).value]:
                somme[sheetTotal.cell(row=i, column=2).value] = item_list
    print(sheetdate[date_row])

    """Write the header of the invoice"""
    creat_invoice = xlsxwriter.Workbook('Sheet_' + str(amount) + '-' + str(sheetdate[date_row]) + ': ' + str(row_startNum) + '.' + 'xlsx')
    invoice = creat_invoice.add_worksheet()
    facthead = "&C3 D'S DISCOUNT STORE"
    invoice.set_header(facthead)
    bold = creat_invoice.add_format({'bold': True})
    money = creat_invoice.add_format({'num_format': '$#,##0.00'})
    invoice.set_column('B:B', 30)
    invoice.set_column('E:E', 12)
    invoice.set_column('F:F', 10)
    invoice.write(6, 0, 'Item ', bold)
    invoice.write(6, 1, 'Description', bold)
    invoice.write(6, 2, 'Qty', bold)
    invoice.write(6, 3, 'Unit Price', bold)
    invoice.write(6, 4, 'Discout', bold)
    invoice.write(6, 5, 'Price', bold)
    invoice.set_column('A1:B5', 15)
    invoice.set_column('A:A', 5)
    invoice.set_column('C:C', 5)
    invoice.set_column('E:E', 15)
    border1 = creat_invoice.add_format({'border': 2})
    blue = creat_invoice.add_format({'color': 'blue'})
    date_format = creat_invoice.add_format({'num_format': 'mm.dd.yyyy', 'color': 'blue'})
    cell_format = creat_invoice.add_format({'align': 'left', 'valign': 'vleft', 'border': '', 'color': 'blue'})
    invoice.merge_range('A1:B5',
                        "3 D'S Discount store\nAddress: 6550 WEST GLENDALE AVE\nGLENDALE ARIZONA, 85301\n",
                        cell_format)
    invoice.write('F1', sheetdate[date_row], date_format)
    invoice.write('E1', 'Invoice Date: ', blue)
    invoice.merge_range('C2:F5',
                        'Phone: 602-518-1743\nFax:\nEmail:LYDIAANTONIO@YAHOO.COM\n', blue)

    """Add Time to the sheet"""
    date_row += 1

    row = 7
    col = 3
    row_1 = 7
    items_list = []
    for items in item_list:
        items_count = item_list.count(items)
        items_list.append((items, items_count))
    item_list_set = list(set(items_list))

    def mult_item(a):
        """merge repeated items"""
        prices = []
        item = []
        number_item = []
        price_number_list = []
        invoice_Tab = []
        price_add = 0
        #price_total = 0
        new_price_number_list = []

        for i, j in a:
           item.append(i[0])
           prices.append(i[1])
           number_item.append(j)
           price_number_list.append((i[1], j))

        for iTem in a:
            if prices.count(iTem[0][1]) > 1:
                new_price_number_list.append(iTem[0][1])
                while iTem[0][1] in prices:
                    prices.remove(iTem[0][1])
                    for price_n in price_number_list:
                        if price_n[0] == iTem[0][1]:
                            price_add+= price_n[1]
                            price_number_list.remove(price_n)
                invoice_Tab.append(((iTem[0][0], iTem[0][1]),  price_add))
                price_add = 0
            else:
                if iTem[0][1] not in new_price_number_list:
                    invoice_Tab.append(iTem)
        return invoice_Tab



    invoice_mult = mult_item(item_list_set)

    def merge_price(n):
        """merge prices that can divided each other"""
        final_inv = []
        ln = 0
        while ln < len(n):
            number = 0
            if n[ln][0][1] != 1:
                for item in n:
                    if item[0][1] != 1 and n[ln] != item:
                        price_max = max(n[ln][0][1], item[0][1]) % min(n[ln][0][1], item[0][1])
                        if price_max == 0 and n[ln][0][1] < item[0][1]:
                            number_item = n[ln][1] + ((item[0][1] * item[1]) / n[ln][0][1])
                            if number_item <= 15:
                                final_inv.append((n[ln][0], number_item))
                                number = n[ln][0][1]
                                n.remove(item)
                                n.remove(n[ln])
                                break
                        elif price_max == 0 and n[ln][0][1] > item[0][1]:
                            number_item = item[1] + ((n[ln][0][1] * n[ln][1]) / item[0][1])
                            if number_item <= 15:
                                number = n[ln][0][1]
                                final_inv.append((item[0], number_item))
                                n.remove(n[ln])
                                n.remove(item)
                                break

            if number == 0:
                final_inv.append(n[ln])
                ln += 1
            else:
                ln += 0
        return (final_inv)

    final_invoice = merge_price(invoice_mult)

    """Writing on Sheet"""
    for it in final_invoice:
        invoice.write(row, 2, it[1])
        invoice.write(row, 1, it[0][0])
        invoice.write(row, col, it[0][1], money)
        invoice.write(row_1, 0, row_1 - 6)
        row += 1
        row_1 += 1
    k = len(final_invoice)
    a = 3
    row += 3
    foot_inv = ['Invoice Subtotal', 'Tax Rate', 'Sales Tax', 'Other', 'Deposit Receive', 'TOTAL']
    for t in foot_inv:
        invoice.write(row, 4, t, bold)
        row += 1
    """Add formula in the invoice sheet"""
    invoice.write_formula(row - 1, 5, '{=SUM(C8:C' + str(k + 7) + '*D8:D' + str(k + 7) + ')}', money)
    for p in range(7, k + 7):
        invoice.write_formula(p, 5, '{=C' + str(p + 1) + '*D' + str(p + 1) + '}', money)

    row_startNum += 1
    r += 1
    n_row += 1
    creat_invoice.close()
